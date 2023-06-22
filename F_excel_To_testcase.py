import openpyxl

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment


#execel地址
excel_read = r'C:\Users\kanli\Downloads\SWC_Interface_PX2_20230620.xlsx'

# 打开第一个Excel文件
workbook1 = openpyxl.load_workbook(excel_read)
workbook2 = Workbook()  # 创建一个工作簿
for k in range(2,len(workbook1.worksheets)):
    sheet_read =  str(workbook1.worksheets[k])[12:-2]
    #workbook1.worksheets[k] 返回值为<Worksheet "xxxxx">,需要去除<Worksheet "  和   ">
    worksheet1 = workbook1[sheet_read]
    worksheet2 = workbook2.create_sheet(sheet_read)#创建一个sheet，放在最后

    worksheet2['A1'] = 'Title'#给单元格赋值
    worksheet2['B1'] = 'Description'#给单元格赋值
    worksheet2['C1'] = 'Step'#给单元格赋值
    worksheet2['D1'] = 'Step Description'#给单元格赋值.
    worksheet2['E1'] = 'Excepted Result'#给单元格赋值

    colA = worksheet1['A']
    # 读取第一个Excel文件的三行，将它们写入第二个Excel文件的三列中
    for i in range(1, len(colA)):

        #testcase_rols
        j1=i*5-3
        j2=i*5-2
        j3=i*5-1
        j4=i*5
        j5=i*5+1

        # 读取第一个Excel文件的第i行三个单元格的值
        source_value = str(worksheet1.cell(row=i+1, column=3).value)
        expection_value = str(worksheet1.cell(row=i+1, column=5).value)
        Debug_Switch = str(worksheet1.cell(row=i+1, column=8).value)
        Debug_Value = str(worksheet1.cell(row=i+1, column=9).value)


        #第一列
        cell_testcase = worksheet2.cell(row=j1, column=1, value= source_value)
        worksheet2.merge_cells(start_row=j1,start_column=1,end_row=j4,end_column=1)#合并单元格
        cell_testcase.alignment = Alignment(horizontal='center', vertical='center')

        #第二列
        cell_testdescription = worksheet2.cell(row=j1, column=2, value= source_value)
        worksheet2.merge_cells(start_row=j1,start_column=2,end_row=j4,end_column=2)#合并单元格
        cell_testdescription.alignment = Alignment(horizontal='center', vertical='center')

        #第三列
        worksheet2.cell(row=j1, column=3, value='1')
        worksheet2.cell(row=j2, column=3, value='2')
        worksheet2.cell(row=j3, column=3, value='3')
        worksheet2.cell(row=j4, column=3, value='4')

        #第四列
        worksheet2.cell(row=j1, column=4, value='set '+ Debug_Switch + '=1')
        worksheet2.cell(row=j2, column=4, value='set '+ Debug_Value + '=1')
        worksheet2.cell(row=j3, column=4, value='Get\n' + source_value + '\n' + expection_value)
        worksheet2.cell(row=j4, column=4, value='Compare\n' + source_value  + '\nto\n' + expection_value)
        worksheet2.cell(row=j5, column=4, value='')

        #第五列
        cell_expection = source_value + '\n = \n' + Debug_Value+'\n = \n' + expection_value
        worksheet2.cell(row=i*5-1, column=5, value = cell_expection)



# 保存第二个Excel文件
workbook2.save('C:/WorkSpace/py_excel/file2.xlsx')

