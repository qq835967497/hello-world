## 模块一：获取.c文件列表
import os
import openpyxl

def get_c_files_path(dir_path):
    """
    获取文件夹下的.c文件的文件名和所在文件夹路径，并将其存入列表中
    @param dir_path: 文件夹路径
    @return: 文件路径列表和文件夹路径列表
    """
    c_files_path_list = []  # 文件路径列表
    parent_path_list = []  # 文件夹路径列表
    for root, dirs, files in os.walk(dir_path):
        for file_name in files:
            if os.path.splitext(file_name)[1] == '.c':  # 判断文件扩展名是否为.c
                file_path = os.path.join(root, file_name).replace('\\', '/')  # 转换为Unix路径
                c_files_path_list.append(file_path)
                parent_path = os.path.abspath(os.path.join(file_path, "..")).replace('\\', '/')  # 获得.c文件上一级目录路径
                parent_path_list.append(parent_path)
    return c_files_path_list, parent_path_list


## 模块二：写入Excel表格
def write_to_excel(workbook_path, sheet_name, file_path_list, parent_path_list):
    """
    新建Excel工作表并将文件名和文件夹路径写入
    @param workbook_path: Excel文件路径
    @param sheet_name: 工作表名称
    @param file_path_list: 文件路径列表
    @param parent_path_list: 文件夹路径列表
    """
    # 初始化workbook
    workbook = openpyxl.Workbook()
    # 创建工作表
    sheet = workbook.create_sheet(sheet_name, 0)

    # 写入表头
    sheet.cell(1, 1).value = '文件名'
    sheet.cell(1, 2).value = '文件夹路径'

    # 写入.c文件名和所在文件夹的路径
    for i in range(len(file_path_list)):
        file_path = file_path_list[i]
        file_name = os.path.basename(file_path)
        parent_path = parent_path_list[i]
        sheet.cell(i+2, 1).value = file_name
        sheet.cell(i+2, 2).value = parent_path

    # 保存Excel文件
    workbook.save(workbook_path)




## 模块三：主要程序逻辑
def main():
    dir_path = 'D:/software'  # 文件夹路径
    workbook_path = 'D:/result.xlsx'  # Excel文件路径
    sheet_name = 'sheet1'  # 工作表名称

    # 获取.c文件路径和所在文件夹列表
    file_path_list, parent_path_list = get_c_files_path(dir_path)

    # 将路径列表写入Excel表格
    write_to_excel(workbook_path, sheet_name, file_path_list, parent_path_list)

if __name__ == "__main__":
    main()