from openpyxl import Workbook
from openpyxl import load_workbook







#1.表格
wb = Workbook()#创建一个表格
wb = load_workbook('excel.xlsx')#加载一个表格
wb.title = 'New Title'#修改表格名称
wb.save('excel.xlsx')#保存表格

print(wb.title)#获取表格的名称



#2.sheet
ws = wb.create_sheet('Mysheet')#创建一个sheet，放在最后
ws = wb.create_sheet('Mysheet1',0)#创建一个sheet，放在最前
ws = wb.create_sheet('Mysheet2',1)#创建一个sheet，放在第二个位置
ws = wb['New Title']#通过名称获取sheet
ws = wb.active#获取当前活跃的sheet
ws.title = 'New Title'#修改sheet的名称

print(wb.sheetnames)#获取所有sheet的名称
print(ws.title)#获取sheet的名称




#3.单元格
ws['A1'] = 42#给单元格赋值
cell = ws['A1']#获取单元格
cell = ws.cell(1,2,10)#给单元格赋值,row=1,column=2,value=10
cell = ws.cell(1,2)#获取单元格,row=1,column=2

print(ws['A1'].value)#获取单元格的值
print(cell.value)#获取单元格的值
print(cell.row)#获取单元格的行
print(cell.column)#获取单元格的列
print(cell.coordinate)#获取单元格的坐标


#4.行和列
colC = ws['C']#获取第C列的所有单元格
col_range = ws['C:D']#获取第C列和第D列的所有单元格
row10 = ws[10]#获取第10行的所有单元格
row_range = ws[5:10]#获取第5行到第10行的所有单元格
for col in ws.iter_cols(min_row=1,max_col=3,max_row=2):#获取第1行到第2行，第1列到第3列的所有单元格
    for cell in col:
        print(cell)

for row in ws.iter_rows(min_row=1,max_col=3,max_row=2):
    for cell in row:
        print(cell)


#5.合并单元格
ws.merge_cells('A2:D2')#合并单元格
ws.unmerge_cells('A2:D2')#取消合并单元格
ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=4)#合并单元格
ws.unmerge_cells(start_row=2,start_column=1,end_row=2,end_column=4)#取消合并单元格


#6.插入行和列
ws.insert_rows(7)#在第7行插入一行
ws.insert_rows(7,5)#在第7行插入5行
ws.insert_cols(7)#在第7列插入一列
ws.insert_cols(7,5)#在第7列插入5列


#7.删除行和列
ws.delete_rows(7)#删除第7行
ws.delete_rows(7,5)#删除第7行开始的5行
ws.delete_cols(7)#删除第7列
ws.delete_cols(7,5)#删除第7列开始的5列


#8.移动行和列
ws.move_range('A1:B5',rows=2,cols=2)#将A1:B5的单元格移动到C3:D7的位置
ws.move_range('A1:B5',rows=2,cols=2,translate=True)#将A1:B5的单元格移动到C3:D7的位置,并且移动单元格的值



#9.公式
ws['A1'].hyperlink#获取单元格的超链接
ws['A1'].hyperlink = 'http://www.baidu.com'#给单元格赋值




#10.图像
from openpyxl.drawing.image import Image
img = Image('logo.png')#创建一个图片
ws.add_image(Image('logo.png'), 'A1')#在A1单元格插入图片
ws.add_image(Image('logo.png'), 'A1:D5')#在A1:D5单元格插入图片

#10.1图片大小与表格大小一致
wb = Workbook()#创建一个表格
ws = wb.active
img = Image(r'D:/workspace/Python_Excel/a.png')#创建一个图片
ws.add_image(img, 'A1')#在A1单元格插入图片
#14.1图片大小与表格大小一致
cell = ws['A1']# 获取图像所在的单元格
column_width = ws.column_dimensions[cell.column_letter].width  # 获取单元格的宽度
img.width = column_width  # 图像宽度与单元格宽度相同
row_height = ws.row_dimensions[cell.row].height # 获取单元格的高度
img.height = row_height  # 图像高度与单元格高度相同
img.anchor = 'absolute'# 图像位置与单元格位置相同












